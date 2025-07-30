import os
import json
from openpyxl import load_workbook

# mapping data types -> PostgreSQL format
type_mapping = {
    "string": "text",
    "boolean": "boolean",
    "double": "double precision",
    "integer": "integer"
}

def run_to_json():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.abspath(os.path.join(current_dir, ".."))

    # หาไฟล์ .xlsx ทั้งหมด
    xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith(".xlsx")]

    if not xlsx_files:
        print("❌ ไม่พบไฟล์ .xlsx ในโฟลเดอร์")
        return

    # แสดงรายการให้เลือก
    print("\n📄 พบไฟล์ .xlsx ดังต่อไปนี้:")
    for idx, file in enumerate(xlsx_files, 1):
        print(f"{idx}. {file}")
    print("พิมพ์ 'back' เพื่อย้อนกลับ")

    while True:
        choice = input("เลือกหมายเลขไฟล์: ").strip().lower()
        if choice == "back":
            return
        if not choice.isdigit() or not (1 <= int(choice) <= len(xlsx_files)):
            print("❌ เลือกหมายเลขที่อยู่ในรายการ")
            continue

        selected_file = xlsx_files[int(choice) - 1]
        break

    file_stem = os.path.splitext(selected_file)[0]
    file_path = os.path.join(base_dir, selected_file)
    output_dir = os.path.join(base_dir, file_stem)
    os.makedirs(output_dir, exist_ok=True)

    print(f"\n🔎 กำลังเปิดไฟล์: {file_path}")

    try:
        wb = load_workbook(filename=file_path, data_only=True)

        # กรอง sheet และแสดงให้เลือก
        valid_sheets = []
        print("\n📝 รายชื่อ Sheet ที่พบ:")
        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            if sheet_name.strip().lower() == "detail":
                continue
            print(f"{idx}. {sheet_name}")
            valid_sheets.append((idx, sheet_name))

        if not valid_sheets:
            print("❌ ไม่พบ Sheet(xlsx) ที่สามารถประมวลผลได้")
            return

        print("\nเลือก Sheet ที่ต้องการสร้าง JSON:")
        sheet_selection = input("พิมพ์หมายเลขหรือ 'all': ").strip().lower()

        if sheet_selection == 'all':
            selected_sheets = [name for _, name in valid_sheets]
        else:
            selected_indices = set(sheet_selection.split(','))
            selected_sheets = []
            for idx_str in selected_indices:
                if idx_str.strip().isdigit():
                    idx = int(idx_str.strip())
                    sheet_tuple = next((s for s in valid_sheets if s[0] == idx), None)
                    if sheet_tuple:
                        selected_sheets.append(sheet_tuple[1])
            if not selected_sheets:
                print("❌ ไม่พบหมายเลขที่ถูกต้อง")
                return

        # ยืนยันข้อมูลก่อนสร้าง
        print("\n📌 คุณเลือกสร้าง JSON จาก Sheet ต่อไปนี้:")
        for name in selected_sheets:
            print(f" - {name}")
        confirm = input("พิมพ์ 'Y' เพื่อยืนยันการสร้างไฟล์: ").strip().lower()
        if confirm != 'y':
            print("❎ ยกเลิกการสร้างไฟล์")
            return

        # เริ่มสร้าง JSON
        for sheet_name in selected_sheets:
            sheet = wb[sheet_name]

            tag_group_name = sheet["D3"].value or sheet_name
            description = sheet["D5"].value or ""

            mappings = []
            row = 12  # เริ่มจาก D12, E12, F12
            position = 1

            while True:
                field_name = sheet[f"D{row}"].value
                data_type = sheet[f"E{row}"].value
                is_identity = sheet[f"F{row}"].value

                if field_name is None and data_type is None and is_identity is None:
                    break

                data_type_lower = (data_type or "").strip().lower()
                pg_data_type = type_mapping.get(data_type_lower, data_type)

                mapping = {
                    "fieldName": field_name,
                    "dataType": pg_data_type,
                    "isNotNull": False,
                    "isIdentity": bool(is_identity),
                    "position": position
                }
                mappings.append(mapping)

                row += 1
                position += 1

            result = {
                "tagGroupName": tag_group_name,
                "description": description,
                "enableHyperTable": False,
                "isView": False,
                "sqlQueryScript": None,
                "tagRelationFieldMappings": mappings
            }

            output_file = os.path.join(output_dir, f"{sheet_name}.json")
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(result, f, indent=4, ensure_ascii=False)

            print(f"✅ บันทึกไฟล์ JSON สำหรับ sheet '{sheet_name}' ที่: {output_file}")

    except Exception as e:
        print(f"\n❌ เกิดข้อผิดพลาด: {e}")
